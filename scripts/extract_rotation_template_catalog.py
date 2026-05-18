#!/usr/bin/env python3
"""Extract the backend rotation workbook into a normalized template catalog.

This is a mechanical extraction step. The source workbook remains outside the
repo; the generated JSON keeps source sheet names so the UI can explain which
manual template matched a staffing request.
"""

from __future__ import annotations

import argparse
import json
import math
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


TIME_RE = re.compile(r"^\s*(\d{1,2})(?::(\d{2}))?\s*([AP]M)?\s*$", re.I)
PERSON_RE = re.compile(r"(\d+(?:\.\d+)?)\s*(?:PERSON|PERSONS|PPL|PEOPLE)", re.I)


def clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").strip()
    return re.sub(r"\s+", " ", text)


def compact_key(value: str) -> str:
    key = re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")
    return key or "lane"


def normalize_time(value: Any) -> str | None:
    if value is None:
        return None
    if hasattr(value, "hour") and hasattr(value, "minute"):
        return f"{int(value.hour):02d}:{int(value.minute):02d}"
    text = clean(value)
    if not text:
        return None
    match = TIME_RE.match(text)
    if not match:
        return None
    hour = int(match.group(1))
    minute = int(match.group(2) or 0)
    suffix = (match.group(3) or "").upper()
    if suffix == "PM" and hour != 12:
        hour += 12
    if suffix == "AM" and hour == 12:
        hour = 0
    if hour > 23 or minute > 59:
        return None
    return f"{hour:02d}:{minute:02d}"


def infer_shift(sheet_name: str, first_time: str | None) -> str:
    name = sheet_name.upper()
    if re.search(r"\bAM\b", name):
        return "AM"
    if re.search(r"\bPM\b", name):
        return "PM"
    if first_time:
        return "AM" if first_time < "12:00" else "PM"
    return "unknown"


def infer_day_type(sheet_name: str) -> str:
    name = sheet_name.upper()
    if "WEEKEND" in name:
        return "weekend"
    if "WEEKDAY" in name:
        return "weekday"
    if any(flag in name for flag in ["HOLIDAY", "EASTER", "CHRISTMAS", "SNOW"]):
        return "special"
    return "unknown"


def infer_person_count(sheet_name: str) -> float | None:
    match = PERSON_RE.search(sheet_name)
    if not match:
        return None
    value = float(match.group(1))
    return int(value) if value.is_integer() else value


def infer_flags(sheet_name: str) -> dict[str, bool]:
    name = sheet_name.upper()
    return {
        "pod": "POD" in name,
        "podPass": "POD PASS" in name,
        "privatePlay": bool(re.search(r"\bPP\b", name)),
        "solo": "SOLO" in name,
        "snow": "SNOW" in name,
        "holiday": "HOLIDAY" in name or "EASTER" in name or "CHRISTMAS" in name,
        "trainee": "TRAINEE" in name,
        "weekend": "WEEKEND" in name,
        "weekday": "WEEKDAY" in name,
    }


def infer_position(header: str) -> str:
    normalized = header.lower()
    if "supervisor" in normalized or normalized in {"sup", "lead"}:
        return "supervisor"
    if "mod" in normalized or "manager" in normalized:
        return "manager"
    if "csr" in normalized or "front" in normalized:
        return "csr"
    return "pct"


def infer_task_key(value: str) -> str:
    text = value.lower()
    compact = re.sub(r"\s+", "", text)
    if not text:
        return "empty"
    if "large" in text and ("daycare" in compact or "day care" in text):
        return "lgdc"
    if "small" in text and ("daycare" in compact or "day care" in text):
        return "smdc"
    if "private play" in text or re.search(r"\bpp\b", text):
        return "pp"
    if "break" in text:
        return "break"
    if "bath" in text:
        return "bath"
    if "transport" in text:
        return "transport"
    if "feeding report" in text:
        return "feeding_report"
    if "feed" in text or "med" in text:
        return "feed"
    if "let" in text or "pod" in text or "opening" in text:
        return "opening"
    if "room mess" in text or "room setup" in text or "room set" in text or "clean beds" in text or "grate" in text:
        return "room_clean"
    if "disinfect" in text:
        return "disinfect"
    if "foam" in text:
        return "foam"
    if "laundry" in text or "dishes" in text or "housekeep" in text:
        return "housekeeping"
    if "eod" in text or "closing" in text or "close" in text:
        return "eod"
    if "lobby" in text:
        return "lobby"
    if "report" in text or "paperwork" in text or "admin" in text:
        return "admin"
    return "float"


def find_header_row(rows: list[list[Any]]) -> int | None:
    best_index = None
    best_score = 0
    for index, row in enumerate(rows[:18]):
        values = [clean(value) for value in row]
        score = 0
        if any(value.lower() == "time" for value in values):
            score += 4
        score += sum(1 for value in values if re.search(r"pct|person|supervisor|mod|manager|csr", value, re.I))
        if score > best_score:
            best_index = index
            best_score = score
    if best_score >= 2:
        return best_index

    # Some copied sheets have no explicit headers; they start with a date row,
    # a "Name:" row, then time rows. Treat the name row as a soft header.
    for index, row in enumerate(rows[:18]):
        values = [clean(value).lower() for value in row]
        if any(value == "name:" or value == "name" for value in values):
            return index
    return None


def infer_time_column(rows: list[list[Any]], header_index: int, header: list[str]) -> int:
    explicit = next((idx for idx, value in enumerate(header) if value.lower() == "time"), None)
    if explicit is not None:
        return explicit
    width = max((len(row) for row in rows[header_index : header_index + 18]), default=0)
    best_col = 0
    best_count = 0
    for col_idx in range(width):
        count = 0
        for row in rows[header_index + 1 : header_index + 16]:
            if col_idx < len(row) and normalize_time(row[col_idx]):
                count += 1
        if count > best_count:
            best_col = col_idx
            best_count = count
    return best_col


def build_merged_value_lookup(ws) -> dict[tuple[int, int], Any]:
    merged_values: dict[tuple[int, int], Any] = {}
    for merged_range in ws.merged_cells.ranges:
        value = ws.cell(merged_range.min_row, merged_range.min_col).value
        if value is None:
            continue
        for row_idx in range(merged_range.min_row, merged_range.max_row + 1):
            for col_idx in range(merged_range.min_col, merged_range.max_col + 1):
                merged_values[(row_idx, col_idx)] = value
    return merged_values


def worksheet_values_with_merged_cells(ws) -> list[list[Any]]:
    merged_values = build_merged_value_lookup(ws)
    rows: list[list[Any]] = []
    for row_idx in range(1, ws.max_row + 1):
        row: list[Any] = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            row.append(cell.value if cell.value is not None else merged_values.get((row_idx, col_idx)))
        rows.append(row)
    return rows


def extract_template(ws, sheet_index: int) -> dict[str, Any] | None:
    rows = worksheet_values_with_merged_cells(ws)
    if not rows:
        return None
    if not any(clean(cell) for row in rows for cell in row):
        return None

    header_index = find_header_row(rows)
    if header_index is None:
        return None

    header = [clean(value) for value in rows[header_index]]
    time_col = infer_time_column(rows, header_index, header)
    lane_columns: list[tuple[int, str]] = []
    for col_idx, value in enumerate(header):
        if col_idx == time_col:
            continue
        if re.search(r"pct|person|supervisor|mod|manager|csr", value, re.I):
            lane_columns.append((col_idx, value))

    if not lane_columns:
        first_time_row = next(
            (
                row
                for row in rows[header_index + 1 :]
                if time_col < len(row) and normalize_time(row[time_col])
            ),
            None,
        )
        if first_time_row:
            for col_idx, value in enumerate(first_time_row):
                if col_idx == time_col or not clean(value):
                    continue
                lane_columns.append((col_idx, f"PERSON {len(lane_columns) + 1}"))

    first_time = None
    slots: list[dict[str, Any]] = []
    cells: list[dict[str, Any]] = []
    for row_idx, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        time_value = normalize_time(row[time_col] if time_col < len(row) else None)
        if not time_value:
            continue
        if first_time is None:
            first_time = time_value
        slots.append({
            "time": time_value,
            "sourceRow": row_idx,
        })
        for col_idx, header_label in lane_columns:
            raw = clean(row[col_idx] if col_idx < len(row) else None)
            if not raw:
                continue
            cells.append({
                "time": time_value,
                "laneKey": f"{compact_key(header_label)}_{col_idx + 1}",
                "raw": raw,
                "taskKey": infer_task_key(raw),
            })

    if not slots and not cells:
        return None

    lanes = [
        {
            "key": f"{compact_key(label)}_{col_idx + 1}",
            "sourceColumn": col_idx + 1,
            "label": label,
            "role": infer_position(label),
        }
        for col_idx, label in lane_columns
    ]
    pct_lane_count = sum(1 for lane in lanes if lane["role"] == "pct")
    support_lanes = sorted({lane["role"] for lane in lanes if lane["role"] != "pct"})

    return {
        "id": compact_key(f"{sheet_index:03d}_{ws.title}"),
        "sourceSheetName": ws.title,
        "sheetIndex": sheet_index,
        "shift": infer_shift(ws.title, first_time),
        "dayType": infer_day_type(ws.title),
        "personCount": infer_person_count(ws.title),
        "pctLaneCount": pct_lane_count,
        "supportRoles": support_lanes,
        "flags": infer_flags(ws.title),
        "lanes": lanes,
        "timeSlots": slots,
        "cells": cells,
        "taskCounts": task_counts(cells),
        "metadata": {
            "headerRow": header_index + 1,
            "firstTime": first_time,
            "lastTime": slots[-1]["time"] if slots else None,
            "nonEmptyTaskCells": len(cells),
            "worksheetMaxRow": ws.max_row,
            "worksheetMaxColumn": ws.max_column,
        },
    }


def task_counts(cells: list[dict[str, Any]]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for cell in cells:
        key = str(cell.get("taskKey") or "float")
        counts[key] = counts.get(key, 0) + 1
    return dict(sorted(counts.items(), key=lambda item: (-item[1], item[0])))


def catalog_summary(templates: list[dict[str, Any]]) -> dict[str, Any]:
    def count_by(key: str) -> dict[str, int]:
        counts: dict[str, int] = {}
        for template in templates:
            value = template.get(key)
            label = "unknown" if value is None else str(value)
            counts[label] = counts.get(label, 0) + 1
        return dict(sorted(counts.items()))

    flags: dict[str, int] = {}
    for template in templates:
        for key, enabled in (template.get("flags") or {}).items():
            if enabled:
                flags[key] = flags.get(key, 0) + 1

    return {
        "templateCount": len(templates),
        "shiftCounts": count_by("shift"),
        "dayTypeCounts": count_by("dayType"),
        "flagCounts": dict(sorted(flags.items())),
        "personCounts": count_by("personCount"),
    }


def make_json_safe(value: Any) -> Any:
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None
    if isinstance(value, dict):
        return {key: make_json_safe(item) for key, item in value.items()}
    if isinstance(value, list):
        return [make_json_safe(item) for item in value]
    return value


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()

    wb = load_workbook(args.workbook, data_only=False, read_only=False)
    templates = []
    skipped = []
    for index, ws in enumerate(wb.worksheets, start=1):
        template = extract_template(ws, index)
        if template:
            templates.append(template)
        else:
            skipped.append({"sheetIndex": index, "sourceSheetName": ws.title})

    catalog = {
        "source": {
            "workbookName": args.workbook.name,
            "workbookPathNote": "Original workbook is intentionally not committed.",
            "extractedBy": "scripts/extract_rotation_template_catalog.py",
        },
        "summary": catalog_summary(templates),
        "skippedSheets": skipped,
        "templates": templates,
    }

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(make_json_safe(catalog), indent=2, sort_keys=True) + "\n", encoding="utf-8")
    print(f"Wrote {len(templates)} templates to {args.output}")
    if skipped:
        print(f"Skipped {len(skipped)} empty/unreadable sheets")


if __name__ == "__main__":
    main()
