#!/usr/bin/env python3
"""Build deterministic Enterprise Directory import data from Skyler's source workbooks."""

from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover - environment guard
    raise SystemExit("openpyxl is required to parse the directory workbooks") from exc


CORPORATE_DEFAULT = Path("/<path>/Downloads/LPHI Directory.xlsx")
RESORTS_DEFAULT = Path("/<path>/Downloads/Gm Contact List (1).xlsx")

PERSON_ALIASES = {
    "hayden.ellison42@example.com": "aubrey-nolan-28",
    "remy.sloan36@example.com": "rowan-beckett-25",
    "avery.foster81@example.com": "cameron-rhodes-15",
    "marlowe.rhodes84@example.com": "logan-hale-1",
    "sage.foster98@example.com": "finley-prescott-21",
    "hayden.rhodes56@example.com": "cameron-marsh-11",
    "kevin.russell@lphik9.com": "cameron-marsh-11",
    "skyler.bennett51@example.com": "ellis-bennett-8",
    "jordan.marsh32@example.com": "avery-bennett-26",
}

DIRECT_TO_SEAN_NOTE = ""


def clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    text = unicodedata.normalize("NFKC", text)
    text = text.replace("\u202c", "")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def slugify(value: str) -> str:
    text = unicodedata.normalize("NFKD", clean(value)).encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^a-zA-Z0-9]+", "-", text.lower()).strip("-")
    return text or "unknown"


def normalize_email(value: Any) -> str:
    email = clean(value).lower()
    if not email:
        return ""
    email = email.replace(" ", "")
    if email.endswith("@lphik9"):
        email += ".com"
    return email


def normalize_phone(value: Any) -> str:
    raw = clean(value)
    if not raw:
        return ""
    digits = re.sub(r"\D", "", raw)
    if len(digits) == 11 and digits.startswith("1"):
        digits = digits[1:]
    if len(digits) == 10:
        return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
    return raw


def split_name(full_name: str) -> tuple[str, str]:
    parts = clean(full_name).split()
    if not parts:
        return "", ""
    if len(parts) == 1:
        return parts[0], ""
    return parts[0], " ".join(parts[1:])


def person_key(first_name: str, last_name: str, email: str = "") -> str:
    normalized_email = normalize_email(email)
    if normalized_email in PERSON_ALIASES:
        return PERSON_ALIASES[normalized_email]
    return slugify(f"{first_name} {last_name}" if first_name or last_name else normalized_email)


def display_resort_name(workbook_name: str) -> str:
    name = clean(workbook_name)
    name = re.sub(r"^K9 Resort\s+", "", name, flags=re.I)
    name = re.sub(r"^K9\s+", "", name, flags=re.I)
    return clean(name)


def parse_city_state_zip(value: str) -> tuple[str, str, str]:
    text = clean(value)
    match = re.match(r"^(?P<city>.*?),\s*(?P<state>[A-Z]{2})\s*(?P<zip>\d{5})?", text)
    if not match:
        return "", "", ""
    return clean(match.group("city")), clean(match.group("state")), clean(match.group("zip"))


def add_person(people: dict[str, dict[str, Any]], first: str, last: str, email: str, phone: str, title: str, source: str, metadata: dict[str, Any]) -> str:
    email = normalize_email(email)
    key = person_key(first, last, email)
    display_name = clean(f"{first} {last}")
    existing = people.get(key, {})
    source_systems = sorted(set(existing.get("source_systems", [])) | {source})
    merged_metadata = {**existing.get("source_metadata", {}), source: metadata}
    people[key] = {
        "person_key": key,
        "first_name": clean(first) or existing.get("first_name", ""),
        "last_name": clean(last) or existing.get("last_name", ""),
        "display_name": display_name or existing.get("display_name", ""),
        "email": email or existing.get("email", ""),
        "work_phone": normalize_phone(phone) or existing.get("work_phone", ""),
        "title": clean(title) or existing.get("title", ""),
        "person_type": existing.get("person_type", "person"),
        "directory_status": "active",
        "source_systems": source_systems,
        "source_metadata": merged_metadata,
    }
    return key


def parse_corporate(path: Path, people: dict[str, dict[str, Any]]) -> None:
    ws = openpyxl.load_workbook(path, data_only=True).active
    headers = [clean(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        values = dict(zip(headers, row))
        first = clean(values.get("First Name"))
        last = clean(values.get("Last Name"))
        if not first and not last:
            continue
        add_person(
            people,
            first,
            last,
            values.get("Email"),
            values.get("Work Phone"),
            values.get("Position"),
            "lphi_directory",
            {"workbook": path.name, "sheet": ws.title, "row": row_number},
        )


def parse_resorts(path: Path, people: dict[str, dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], set[str]]:
    ws = openpyxl.load_workbook(path, data_only=True).active
    resort_blocks = [row for row in range(1, ws.max_row + 1) if clean(ws.cell(row, 1).value).lower() == "resort"]
    locations: list[dict[str, Any]] = []
    assignments: list[dict[str, Any]] = []
    gaps: list[dict[str, Any]] = []
    regional_keys: set[str] = set()

    for block_index, start_row in enumerate(resort_blocks):
        end_row = (resort_blocks[block_index + 1] - 1) if block_index + 1 < len(resort_blocks) else ws.max_row
        labels: dict[str, int] = {}
        for row in range(start_row, end_row + 1):
            label = clean(ws.cell(row, 1).value).lower()
            if label:
                labels[label] = row

        resort_cols = [
            col for col in range(3, ws.max_column + 1, 2)
            if clean(ws.cell(start_row, col).value)
        ]
        for col in resort_cols:
            raw_name = clean(ws.cell(start_row, col).value)
            display_name = display_resort_name(raw_name)
            key = slugify(display_name)
            address_1 = clean(ws.cell(labels.get("address", start_row + 1), col).value)
            second_address_row = labels.get("address", start_row + 1) + 1 if labels.get("address") else start_row + 2
            address_2 = clean(ws.cell(second_address_row, col).value)
            city, state, postal_code = parse_city_state_zip(address_2)
            region = clean(ws.cell(1, col).value) or state
            phone_row = labels.get("resort phone") or labels.get("resort landline")
            email_row = labels.get("e-mail")
            hours_rows = [row for row in range(labels.get("hours", start_row), min(end_row, labels.get("hours", start_row) + 2) + 1)] if labels.get("hours") else []
            hours = [clean(ws.cell(row, col).value) for row in hours_rows if clean(ws.cell(row, col).value)]

            locations.append({
                "location_key": key,
                "source_location_name": raw_name,
                "display_name": display_name,
                "state_code": state,
                "region_label": region,
                "address_line1": address_1,
                "address_line2": address_2,
                "city": city,
                "postal_code": postal_code,
                "resort_phone": normalize_phone(ws.cell(phone_row, col).value) if phone_row else "",
                "resort_email": normalize_email(ws.cell(email_row, col).value) if email_row else "",
                "hours": hours,
                "directory_status": "active",
                "source_metadata": {"workbook": path.name, "sheet": ws.title, "block_start_row": start_row, "column": col},
            })

            gm_name = clean(ws.cell(labels.get("gm", 0), col).value) if labels.get("gm") else ""
            gm_email = normalize_email(ws.cell(labels.get("gm  e-mail", 0), col).value) if labels.get("gm  e-mail") else ""
            gm_phone = normalize_phone(ws.cell(labels.get("gm phone", 0), col).value) if labels.get("gm phone") else ""
            if gm_name or gm_email:
                first, last = split_name(gm_name)
                gm_key = add_person(
                    people,
                    first,
                    last,
                    gm_email,
                    gm_phone,
                    "General Manager",
                    "gm_contact_list",
                    {"workbook": path.name, "sheet": ws.title, "row": labels.get("gm"), "column": col, "resort": raw_name},
                )
                assignments.append({
                    "person_key": gm_key,
                    "location_key": key,
                    "responsibility_type": "general_manager",
                    "title": "General Manager",
                    "source": "gm_contact_list",
                    "source_metadata": {"row": labels.get("gm"), "column": col},
                })
            else:
                gaps.append({
                    "gap_key": f"{key}:gm:missing",
                    "entity_type": "location",
                    "entity_key": key,
                    "location_key": key,
                    "person_key": "",
                    "field_name": "general_manager",
                    "severity": "needs_data",
                    "status_label": "Needs data",
                    "detail": f"{display_name} is missing GM name/email/phone data in the workbook.",
                    "source_metadata": {"workbook": path.name, "sheet": ws.title, "block_start_row": start_row, "column": col},
                })

            regional_name = clean(ws.cell(labels.get("ervp name", 0), col).value) if labels.get("ervp name") else ""
            regional_email = normalize_email(ws.cell(labels.get("ervp e-mail", 0), col).value) if labels.get("ervp e-mail") else ""
            regional_phone = normalize_phone(ws.cell(labels.get("ervp phone", 0), col).value) if labels.get("ervp phone") else ""
            if regional_name or regional_email:
                first, last = split_name(regional_name)
                regional_key = add_person(
                    people,
                    first,
                    last,
                    regional_email,
                    regional_phone,
                    "Regional Manager",
                    "gm_contact_list",
                    {"workbook": path.name, "sheet": ws.title, "row": labels.get("ervp name"), "column": col, "resort": raw_name},
                )
                regional_keys.add(regional_key)
                assignments.append({
                    "person_key": regional_key,
                    "location_key": key,
                    "responsibility_type": "regional_manager",
                    "title": "Regional Manager",
                    "source": "gm_contact_list",
                    "source_metadata": {"row": labels.get("ervp name"), "column": col},
                })
            else:
                gaps.append({
                    "gap_key": f"{key}:regional:missing",
                    "entity_type": "location",
                    "entity_key": key,
                    "location_key": key,
                    "person_key": "",
                    "field_name": "regional_manager",
                    "severity": "needs_data",
                    "status_label": "Needs data",
                    "detail": f"{display_name} is missing regional manager data in the workbook.",
                    "source_metadata": {"workbook": path.name, "sheet": ws.title, "block_start_row": start_row, "column": col},
                })

            if gm_name and regional_name.lower() == "sean powell":
                gaps.append({
                    "gap_key": f"{key}:regional:sean-direct",
                    "entity_type": "location",
                    "entity_key": key,
                    "location_key": key,
                    "person_key": "",
                    "field_name": "regional_manager",
                    "severity": "note",
                    "status_label": "Needs data",
                    "detail": f"{display_name}: {DIRECT_TO_SEAN_NOTE}",
                    "source_metadata": {"workbook": path.name, "sheet": ws.title, "row": labels.get("ervp name"), "column": col},
                })

    # Workbook summary table says CA/AZ regional manager is Elliot, while rows
    # currently point CA/AZ resorts to Emerson. Preserve that contradiction.
    gaps.append({
        "gap_key": "regional-list:ca-az-conflict",
        "entity_type": "regional_assignment",
        "entity_key": "ca-az",
        "location_key": "",
        "person_key": "avery-bennett-26",
        "field_name": "regional_manager",
        "severity": "note",
        "status_label": "Needs data",
        "detail": "",
        "source_metadata": {"workbook": path.name, "sheet": ws.title, "rows": [85, 86, 87, 88]},
    })

    return locations, assignments, gaps, regional_keys


def build_edges(people: dict[str, dict[str, Any]], assignments: list[dict[str, Any]]) -> list[dict[str, Any]]:
    edges: dict[tuple[str, str], dict[str, Any]] = {}

    def add(parent: str, child: str, source: str, primary: bool = True, note: str = "") -> None:
        if not parent or not child or parent == child:
            return
        edges[(parent, child)] = {
            "parent_key": parent,
            "child_key": child,
            "relationship_type": "reports_to",
            "is_primary": primary,
            "source": source,
            "source_metadata": {"note": note} if note else {},
        }

    add("logan-hale-1", "cameron-rhodes-15", "hierarchy_rule", True, "")
    add("finley-prescott-21", "cameron-rhodes-15", "hierarchy_rule", False, "")
    add("cameron-rhodes-15", "rowan-beckett-25", "hierarchy_rule", True)
    add("rowan-beckett-25", "aubrey-nolan-28", "hierarchy_rule", True, "")

    for key in ("cameron-marsh-11", "ellis-bennett-8", "avery-bennett-26"):
        if key in people:
            add("rowan-beckett-25", key, "hierarchy_rule", True)

    regional_by_location: dict[str, str] = {}
    gm_by_location: dict[str, str] = {}
    for assignment in assignments:
        if assignment["responsibility_type"] == "regional_manager":
            regional_by_location[assignment["location_key"]] = assignment["person_key"]
        if assignment["responsibility_type"] == "general_manager":
            gm_by_location[assignment["location_key"]] = assignment["person_key"]

    for location_key, gm_key in gm_by_location.items():
        regional_key = regional_by_location.get(location_key)
        if gm_key == "aubrey-nolan-28":
            continue
        if regional_key == "rowan-beckett-25" or not regional_key:
            add("rowan-beckett-25", gm_key, "gm_contact_list", True, DIRECT_TO_SEAN_NOTE)
        else:
            add(regional_key, gm_key, "gm_contact_list", True)

    return sorted(edges.values(), key=lambda row: (row["parent_key"], row["child_key"]))


def build_dataset(corporate: Path, resorts: Path) -> dict[str, Any]:
    people: dict[str, dict[str, Any]] = {}
    parse_corporate(corporate, people)
    locations, assignments, gaps, _regional_keys = parse_resorts(resorts, people)

    # Skyler is a single directory person with Adair Forsythe responsibility attached.
    zack = people.get("aubrey-nolan-28")
    if zack:
        zack["title"] = "Director of Resorts / General Manager, Adair Forsythe"
        zack["source_metadata"]["hierarchy_rule"] = {
            "note": "Represent Skyler once under Sawyer; attach Adair Forsythe GM responsibility."
        }

    return {
        "source_files": {"corporate": str(corporate), "resorts": str(resorts)},
        "people": sorted(people.values(), key=lambda row: row["display_name"]),
        "locations": sorted(locations, key=lambda row: row["display_name"]),
        "person_locations": sorted(assignments, key=lambda row: (row["location_key"], row["responsibility_type"], row["person_key"])),
        "edges": build_edges(people, assignments),
        "gaps": sorted(gaps, key=lambda row: row["gap_key"]),
    }


def sql_literal(value: Any) -> str:
    if value is None:
        return "NULL"
    return "'" + str(value).replace("'", "''") + "'"


def emit_sql(dataset: dict[str, Any]) -> str:
    payload = json.dumps(dataset, indent=2, sort_keys=True)
    return f"""-- Deterministic import generated by scripts/import_enterprise_directory.py.
-- Source workbooks:
-- - {dataset['source_files']['corporate']}
-- - {dataset['source_files']['resorts']}

WITH payload AS (
  SELECT $k9dir${payload}$k9dir$::jsonb AS data
),
locations_input AS (
  SELECT *
  FROM payload, jsonb_to_recordset(data->'locations') AS row(
    location_key text,
    source_location_name text,
    display_name text,
    state_code text,
    region_label text,
    address_line1 text,
    address_line2 text,
    city text,
    postal_code text,
    resort_phone text,
    resort_email text,
    hours jsonb,
    directory_status text,
    source_metadata jsonb
  )
),
people_input AS (
  SELECT *
  FROM payload, jsonb_to_recordset(data->'people') AS row(
    person_key text,
    first_name text,
    last_name text,
    display_name text,
    email text,
    work_phone text,
    title text,
    person_type text,
    directory_status text,
    source_systems text[],
    source_metadata jsonb
  )
),
person_locations_input AS (
  SELECT *
  FROM payload, jsonb_to_recordset(data->'person_locations') AS row(
    person_key text,
    location_key text,
    responsibility_type text,
    title text,
    source text,
    source_metadata jsonb
  )
),
edges_input AS (
  SELECT *
  FROM payload, jsonb_to_recordset(data->'edges') AS row(
    parent_key text,
    child_key text,
    relationship_type text,
    is_primary boolean,
    source text,
    source_metadata jsonb
  )
),
gaps_input AS (
  SELECT *
  FROM payload, jsonb_to_recordset(data->'gaps') AS row(
    gap_key text,
    entity_type text,
    entity_key text,
    location_key text,
    person_key text,
    field_name text,
    severity text,
    status_label text,
    detail text,
    source_metadata jsonb
  )
),
upsert_locations AS (
  INSERT INTO public.enterprise_directory_locations (
    location_key,
    source_location_name,
    display_name,
    state_code,
    region_label,
    address_line1,
    address_line2,
    city,
    postal_code,
    resort_phone,
    resort_email,
    hours,
    directory_status,
    source_metadata
  )
  SELECT
    location_key,
    source_location_name,
    display_name,
    NULLIF(state_code, ''),
    NULLIF(region_label, ''),
    NULLIF(address_line1, ''),
    NULLIF(address_line2, ''),
    NULLIF(city, ''),
    NULLIF(postal_code, ''),
    NULLIF(resort_phone, ''),
    NULLIF(resort_email, ''),
    COALESCE(hours, '[]'::jsonb),
    directory_status,
    source_metadata
  FROM locations_input
  ON CONFLICT (location_key) DO UPDATE SET
    source_location_name = EXCLUDED.source_location_name,
    display_name = EXCLUDED.display_name,
    state_code = EXCLUDED.state_code,
    region_label = EXCLUDED.region_label,
    address_line1 = EXCLUDED.address_line1,
    address_line2 = EXCLUDED.address_line2,
    city = EXCLUDED.city,
    postal_code = EXCLUDED.postal_code,
    resort_phone = EXCLUDED.resort_phone,
    resort_email = EXCLUDED.resort_email,
    hours = EXCLUDED.hours,
    directory_status = EXCLUDED.directory_status,
    source_metadata = EXCLUDED.source_metadata,
    updated_at = now()
  RETURNING id, location_key
),
upsert_people AS (
  INSERT INTO public.enterprise_directory_people (
    person_key,
    first_name,
    last_name,
    display_name,
    email,
    work_phone,
    title,
    person_type,
    directory_status,
    source_systems,
    source_metadata
  )
  SELECT
    person_key,
    NULLIF(first_name, ''),
    NULLIF(last_name, ''),
    display_name,
    NULLIF(email, ''),
    NULLIF(work_phone, ''),
    NULLIF(title, ''),
    person_type,
    directory_status,
    source_systems,
    source_metadata
  FROM people_input
  ON CONFLICT (person_key) DO UPDATE SET
    first_name = EXCLUDED.first_name,
    last_name = EXCLUDED.last_name,
    display_name = EXCLUDED.display_name,
    email = EXCLUDED.email,
    work_phone = EXCLUDED.work_phone,
    title = EXCLUDED.title,
    person_type = EXCLUDED.person_type,
    directory_status = EXCLUDED.directory_status,
    source_systems = EXCLUDED.source_systems,
    source_metadata = EXCLUDED.source_metadata,
    updated_at = now()
  RETURNING id, person_key
),
delete_assignments AS (
  DELETE FROM public.enterprise_directory_person_locations
  WHERE source = 'gm_contact_list'
),
insert_assignments AS (
  INSERT INTO public.enterprise_directory_person_locations (
    person_id,
    location_id,
    responsibility_type,
    title,
    source,
    source_metadata
  )
  SELECT
    p.id,
    l.id,
    pli.responsibility_type,
    NULLIF(pli.title, ''),
    pli.source,
    pli.source_metadata
  FROM person_locations_input pli
  JOIN public.enterprise_directory_people p ON p.person_key = pli.person_key
  JOIN public.enterprise_directory_locations l ON l.location_key = pli.location_key
  ON CONFLICT (person_id, location_id, responsibility_type) DO UPDATE SET
    title = EXCLUDED.title,
    source = EXCLUDED.source,
    source_metadata = EXCLUDED.source_metadata,
    updated_at = now()
),
delete_edges AS (
  DELETE FROM public.enterprise_directory_edges
  WHERE source IN ('hierarchy_rule', 'gm_contact_list')
),
insert_edges AS (
  INSERT INTO public.enterprise_directory_edges (
    parent_person_id,
    child_person_id,
    relationship_type,
    is_primary,
    source,
    source_metadata
  )
  SELECT
    parent.id,
    child.id,
    ei.relationship_type,
    ei.is_primary,
    ei.source,
    ei.source_metadata
  FROM edges_input ei
  JOIN public.enterprise_directory_people parent ON parent.person_key = ei.parent_key
  JOIN public.enterprise_directory_people child ON child.person_key = ei.child_key
  WHERE parent.id <> child.id
  ON CONFLICT (parent_person_id, child_person_id, relationship_type) DO UPDATE SET
    is_primary = EXCLUDED.is_primary,
    source = EXCLUDED.source,
    source_metadata = EXCLUDED.source_metadata,
    updated_at = now()
),
delete_gaps AS (
  DELETE FROM public.enterprise_directory_data_gaps
  WHERE source = 'directory_workbook_import'
),
insert_gaps AS (
  INSERT INTO public.enterprise_directory_data_gaps (
    gap_key,
    entity_type,
    entity_key,
    location_id,
    person_id,
    field_name,
    severity,
    status_label,
    detail,
    source,
    source_metadata
  )
  SELECT
    gi.gap_key,
    gi.entity_type,
    gi.entity_key,
    l.id,
    p.id,
    gi.field_name,
    gi.severity,
    gi.status_label,
    gi.detail,
    'directory_workbook_import',
    gi.source_metadata
  FROM gaps_input gi
  LEFT JOIN public.enterprise_directory_locations l ON l.location_key = NULLIF(gi.location_key, '')
  LEFT JOIN public.enterprise_directory_people p ON p.person_key = NULLIF(gi.person_key, '')
  ON CONFLICT (gap_key) DO UPDATE SET
    entity_type = EXCLUDED.entity_type,
    entity_key = EXCLUDED.entity_key,
    location_id = EXCLUDED.location_id,
    person_id = EXCLUDED.person_id,
    field_name = EXCLUDED.field_name,
    severity = EXCLUDED.severity,
    status_label = EXCLUDED.status_label,
    detail = EXCLUDED.detail,
    source = EXCLUDED.source,
    source_metadata = EXCLUDED.source_metadata,
    updated_at = now()
)
SELECT
  (SELECT count(*) FROM locations_input) AS imported_locations,
  (SELECT count(*) FROM people_input) AS imported_people,
  (SELECT count(*) FROM person_locations_input) AS imported_person_locations,
  (SELECT count(*) FROM edges_input) AS imported_edges,
  (SELECT count(*) FROM gaps_input) AS imported_gaps;
"""


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--corporate", type=Path, default=CORPORATE_DEFAULT)
    parser.add_argument("--resorts", type=Path, default=RESORTS_DEFAULT)
    parser.add_argument("--format", choices=["json", "sql", "summary"], default="json")
    parser.add_argument("--inject-into", type=Path, help="Replace the generated import section inside a migration file.")
    args = parser.parse_args()

    dataset = build_dataset(args.corporate, args.resorts)
    if args.inject_into:
      seed_sql = emit_sql(dataset).rstrip()
      migration_text = args.inject_into.read_text()
      start_marker = "-- BEGIN GENERATED DIRECTORY IMPORT"
      end_marker = "-- END GENERATED DIRECTORY IMPORT"
      start = migration_text.index(start_marker) + len(start_marker)
      end = migration_text.index(end_marker)
      next_text = migration_text[:start] + "\n" + seed_sql + "\n" + migration_text[end:]
      args.inject_into.write_text(next_text)
      return 0

    if args.format == "sql":
      sys.stdout.write(emit_sql(dataset))
    elif args.format == "summary":
      summary = {
          "people": len(dataset["people"]),
          "locations": len(dataset["locations"]),
          "person_locations": len(dataset["person_locations"]),
          "edges": len(dataset["edges"]),
          "gaps": len(dataset["gaps"]),
          "gap_keys": [gap["gap_key"] for gap in dataset["gaps"]],
      }
      sys.stdout.write(json.dumps(summary, indent=2, sort_keys=True) + "\n")
    else:
      sys.stdout.write(json.dumps(dataset, indent=2, sort_keys=True) + "\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
